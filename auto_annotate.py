# -*- coding: utf-8 -*-
"""
SMS Annotation Script - Annotates all 10000 rows in normal_2w_output.xlsx
Adds: final_label, confidence, rationale columns
Priority: TRANSACTION > AD > HARASS > FRAUD > NEEDS_REVIEW
"""
import re
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

INPUT_FILE = r'C:\Users\woshinibaba\Downloads\normal_2w_output.xlsx'

def classify_sms(text, orig_label):
    """Classify SMS based on content patterns and original label"""
    if not text or not text.strip():
        return 'NEEDS_REVIEW', '低', '短信内容为空'
    
    text_lower = text.lower()
    
    # Priority 1: FRAUD (scam/fraud indicators)
    fraud_keywords = [
        '涉嫌诈骗', '涉嫌恶意透支', '冻结支付账户', '公安', '警察',
        '法院传票', '拘捕令', '涉嫌洗钱', '包裹涉毒', '社保异常',
        '医保异常', '涉嫌贩毒', '逮捕令'
    ]
    for kw in fraud_keywords:
        if kw in text:
            return 'FRAUD', '高', f'包含诈骗关键词: {kw}'
    
    # Priority 2: HARASS (threats, collection, intimidation)
    harass_keywords = [
        '严重逾期', '恶意透支', '强制执行', '起诉', '法庭', '律师函',
        '上报征信', '列入失信', '冻结账户', '划扣工资', '联系单位',
        '户籍地', '亲属', '配偶', '催收', '催缴', '逾期记录',
        '报送人行', '人行金融信用', '征信系统', '不良信用',
        '默认拒偿', '永不减免', '申请司法', '诉讼', '开庭',
        '最后期限', '立即还款', '否则将', '后果自负', '承担法律责任'
    ]
    harass_count = sum(1 for kw in harass_keywords if kw in text)
    if harass_count >= 2:
        return 'HARASS', '高', f'包含多个催收/威胁关键词({harass_count}个)'
    
    # Check for loan collection with account actions
    collection_patterns = [
        r'合同逾期.*全款偿还',
        r'分期资格.*取消',
        r'逾期记录.*提报',
        r'强制执行.*查封',
        r'催缴.*公开',
        r'联系.*亲属',
    ]
    for pat in collection_patterns:
        if re.search(pat, text):
            return 'HARASS', '高', f'匹配催收模式: {pat}'
    
    # Priority 3: TRANSACTION (service notifications, transactions)
    # Bank/financial transactions
    transaction_keywords = [
        '还款', '扣款', '到账', '余额', '账单', '提现', '充值', '消费',
        '收入', '支出', '转账', '汇款', '入账', '支出', '自动还款',
        '还款成功', '还款失败', '信用卡', '银行', '借呗', '花呗',
        '信用卡账单', '本期账单', '未还', '已还', '尾号', '卡号',
        '消费提醒', '支出提醒', '收入提醒', '动账通知',
        '积分', '券', '红包', '电子券', '流量', '话费',
        '退款', '返现', '返还', '到账通知',
        '订单', '取件', '派送', '快递', '包裹', '运单',
        '已发货', '已送达', '配送', '物流',
        '验证码', '验证码为', '登录验证', '安全验证',
        '余额不足', '扣费', '代交', '代扣',
    ]
    trans_count = sum(1 for kw in transaction_keywords if kw in text)
    
    # Strong transaction indicators
    strong_trans_patterns = [
        r'还款.*\d+\.?\d*元',
        r'到账.*\d+\.?\d*元',
        r'余额.*\d+\.?\d*元',
        r'账单.*\d+\.?\d*元',
        r'消费.*\d+\.?\d*元',
        r'充值.*\d+\.?\d*元',
        r'提现.*\d+\.?\d*元',
        r'尾号\w+.*\d+\.?\d*元',
        r'验证码.*\d{4,8}',
        r'\d+\.?\d*元.*还款',
        r'取件码.*\w+',
        r'凭\w+.*领取',
    ]
    strong_trans = any(re.search(pat, text) for pat in strong_trans_patterns)
    
    # Government/public service
    gov_keywords = [
        '中国残联', '助残日', '共青团', '12355', '公益热线',
        '防汛', '抗旱', '应急管理部门', '气象预警', '高温预警',
        '公安', '交警', '消防', '安全提醒', '防诈骗',
        '中国地震', '预警信息', '应急预警',
        '科技工作者日', '科协',
        '教育部', '教育厅', '整治.*办学', '严禁提前开学',
        '存款保险', '中国人民银行',
        '全国科技工作者日', '科协',
    ]
    is_gov = any(kw in text for kw in gov_keywords)
    
    # News/media content (not AD)
    news_keywords = [
        '江西日报', '央视新闻', '新华社', '中新网', '人民日报',
        '大江网', '信息日报', '四川手机报', '山西日报',
        '新闻', '日讯', '日电', '记者.*报道',
    ]
    is_news = any(kw in text for kw in news_keywords)
    
    # Express/delivery patterns
    express_patterns = [
        r'取件码', r'凭\w+.*领取', r'运单尾号', r'包裹.*派送',
        r'快递.*到', r'韵达', r'顺丰', r'中通', r'圆通', r'极兔',
        r'京东配送', r'菜鸟驿站', r'丰巢', r'妈妈驿站',
    ]
    is_express = any(re.search(pat, text) for pat in express_patterns)
    
    # Operator service notifications
    operator_service_patterns = [
        r'流量.*使用', r'语音.*使用', r'余额', r'话费',
        r'中国移动', r'中国联通', r'中国电信',
        r'套餐', r'月租', r'充值', r'消费',
        r'积分.*领取', r'权益.*领取',
    ]
    is_operator = any(re.search(pat, text) for pat in operator_service_patterns)
    
    # Classify based on patterns
    # Government/News → NEEDS_REVIEW
    if is_gov and not any(kw in text for kw in ['权益', '会员', '流量', '优惠', '红包']):
        return 'NEEDS_REVIEW', '高', '政府/公益信息，无商业推广'
    
    if is_news and not any(kw in text for kw in ['权益', '会员', '流量', '优惠', '红包']):
        return 'NEEDS_REVIEW', '高', '新闻媒体信息，无商业推广'
    
    # Express delivery → TRANSACTION
    if is_express and not any(kw in text for kw in ['贷款', '额度', '借款', '审核']):
        return 'TRANSACTION', '高', '快递/物流服务通知'
    
    # Strong transaction indicators
    if strong_trans and trans_count >= 2:
        return 'TRANSACTION', '高', '包含明确的交易信息（金额/账单/到账）'
    
    # Operator services with transactions
    if is_operator and trans_count >= 3:
        return 'TRANSACTION', '高', '运营商服务通知（余额/流量/账单）'
    
    # Personal/business notifications
    personal_patterns = [
        r'您好.*客户', r'尊敬的.*用户', r'温馨提示',
        r'您的.*已', r'账户.*变动',
    ]
    if any(re.search(pat, text) for pat in personal_patterns):
        if trans_count >= 2:
            return 'TRANSACTION', '中', '个人/账户服务通知'
    
    # Priority 4: AD (advertisements)
    ad_keywords = [
        '贷款', '额度', '借款', '预授信', '预放', '预审',
        '点击.*领取', '点击.*查看', '点击.*查询', '点击.*参与',
        '拒收请回复R', '拒收请回复 R',
        '会员', '权益', '优惠', '红包', '福利',
        '限时', '免费领取', '免费', '特价',
        '恭喜.*获得', '恭喜.*中奖',
        '预授信', '预批', '特批', '放宽',
        '注册.*送', '邀请.*体验',
        '咨询电话', '客服电话', '详询.*\d{3,}',
        'http', 'https', '链接',
    ]
    ad_count = sum(1 for kw in ad_keywords if kw in text)
    
    # Loan ads (very common)
    loan_keywords = [
        '借款.*已', '额度.*已', '贷款.*已', '预放', '预审',
        '消费额', '信用贷', '贷款额度', '借款额度',
        '激活.*额度', '查看.*额度', '领取.*额度',
        '\d+元.*已', '预汇入', '预转入',
    ]
    is_loan_ad = any(kw in text for kw in loan_keywords)
    
    if is_loan_ad:
        return 'AD', '高', '贷款产品推广广告'
    
    if ad_count >= 3:
        return 'AD', '高', '包含多个广告特征（链接/推广/诱导点击）'
    
    # Medical ads
    medical_ad_keywords = [
        '筛查', '普查', '补助', '援助', '申请.*回',
        '白癜风', '银屑病', '胎记', '癫痫',
        '了解详情', '申请检查',
    ]
    is_medical_ad = any(kw in text for kw in medical_ad_keywords)
    if is_medical_ad:
        return 'AD', '中', '医疗推广信息'
    
    # Education ads
    edu_ad_keywords = [
        '课程', '培训', '资料已发', '教程已发',
        '点击.*领取', '不点.*失效', '速领',
    ]
    is_edu_ad = any(kw in text for kw in edu_ad_keywords)
    if is_edu_ad:
        return 'AD', '中', '教育培训推广'
    
    # Operator marketing (NOT service notification)
    if is_operator:
        # Check if it's purely marketing
        marketing_keywords = [
            '权益', '会员', '优惠', '红包', '福利',
            '限时', '免费领取', '恭喜.*获得',
            '点击.*领取', '点击.*参与', '点击.*查看',
            '抽奖', '赢取', '酬金',
        ]
        is_marketing = any(kw in text for kw in marketing_keywords)
        if is_marketing:
            return 'AD', '中', '运营商营销推广'
    
    # News media with ads
    if is_news and ad_count >= 2:
        return 'NEEDS_REVIEW', '中', '新闻媒体信息，但含推广链接'
    
    # Personal/business
    personal_keywords = [
        '您好', '尊敬的', '温馨提示',
        '会议', '通知', '请.*回复',
    ]
    is_personal = any(kw in text for kw in personal_keywords)
    if is_personal and not is_express and not is_operator:
        return 'NEEDS_REVIEW', '低', '个人/商务通知，需人工判断'
    
    # Default
    return 'NEEDS_REVIEW', '低', '无法明确归类，需人工判断'


def classify_row(row_text, orig_label):
    """Enhanced classification with context from original label"""
    text = str(row_text) if row_text else ''
    
    # Use original label as hint but re-classify based on content
    if orig_label == '黑短信':
        # Check if it's actually a valid service message mislabeled as black
        valid_keywords = ['快递', '包裹', '取件', '验证码', '还款', '账单', '到账']
        if any(kw in text for kw in valid_keywords):
            label, conf, reason = classify_sms(text, orig_label)
            if label in ['TRANSACTION', 'NEEDS_REVIEW']:
                return label, conf, f'原标签为黑短信但内容为正常服务: {reason}'
    
    # For garbled labels, re-classify
    if orig_label and not any(kw in orig_label for kw in ['广告', '正常', '黑']):
        return classify_sms(text, orig_label)
    
    return classify_sms(text, orig_label)


def main():
    print('Loading workbook...')
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    
    # Read all rows
    rows = list(ws.iter_rows(values_only=True))
    print(f'Total rows: {len(rows)}')
    
    # Check if annotation columns already exist
    header = rows[0]
    print(f'Header: {header}')
    
    # Add annotation columns if not present
    if len(header) < 7:
        ws.cell(row=1, column=5, value='final_label')
        ws.cell(row=1, column=6, value='confidence')
        ws.cell(row=1, column=7, value='rationale')
    
    # Annotate all rows
    annotation_count = {'TRANSACTION': 0, 'AD': 0, 'HARASS': 0, 'FRAUD': 0, 'NEEDS_REVIEW': 0}
    
    for i in range(1, len(rows)):
        text = rows[i][0] if rows[i][0] else ''
        orig_label = rows[i][1] if len(rows[i]) > 1 else ''
        
        final_label, confidence, rationale = classify_row(text, orig_label)
        
        # Write to Excel
        ws.cell(row=i+1, column=5, value=final_label)
        ws.cell(row=i+1, column=6, value=confidence)
        ws.cell(row=i+1, column=7, value=rationale)
        
        annotation_count[final_label] += 1
    
    # Save
    print('Saving workbook...')
    wb.save(INPUT_FILE)
    print('Done!')
    
    # Print statistics
    print('\n=== Annotation Statistics ===')
    total = sum(annotation_count.values())
    for label, count in sorted(annotation_count.items()):
        pct = count / total * 100
        print(f'{label}: {count} ({pct:.1f}%)')
    print(f'Total: {total}')


if __name__ == '__main__':
    main()
