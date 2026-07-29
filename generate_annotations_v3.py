# -*- coding: utf-8 -*-
"""Generate complete annotations for all 10000 SMS rows - v3 improved"""
import openpyxl, sys, re
from collections import Counter
sys.stdout.reconfigure(encoding='utf-8')

INPUT_FILE = r'C:\Users\woshinibaba\Downloads\normal_2w_output.xlsx'
OUTPUT_CSV = r'C:\Users\woshinibaba\Documents\oppo的项目\Android_SMS_Classifier\annotations_all_10000.csv'

def classify_row(text, orig_label):
    if not text or not str(text).strip():
        return 'NEEDS_REVIEW', '低', '短信内容为空'
    text = str(text)
    
    # FRAUD
    fraud_kws = ['涉嫌诈骗', '涉嫌恶意透支', '冻结支付账户', '法院传票',
                 '拘捕令', '涉嫌洗钱', '包裹涉毒', '社保异常', '逮捕令']
    if any(kw in text for kw in fraud_kws):
        return 'FRAUD', '高', '包含诈骗关键词'
    
    # HARASS
    harass_kws = ['严重逾期', '恶意透支', '强制执行', '起诉', '法庭', '律师函',
                  '上报征信', '列入失信', '冻结账户', '划扣工资', '联系单位',
                  '户籍地', '催收', '催缴', '逾期记录', '报送人行',
                  '征信系统', '不良信用', '默认拒偿', '永不减免',
                  '申请司法', '诉讼', '开庭', '最后期限', '否则将',
                  '后果自负', '采取.*措施', '查封.*冻结', '拒不还款',
                  '诉前告知', '移交司法', '申请.*强制执行', '仲裁委',
                  '仲裁', '欠款.*公开', '催缴流程']
    if sum(1 for kw in harass_kws if kw in text) >= 1:
        return 'HARASS', '高', '包含催收/威胁关键词'
    
    collection_pats = [r'合同逾期.*全款偿还', r'分期资格.*取消',
                       r'强制执行.*查封', r'催缴.*公开', r'联系.*亲属',
                       r'提交.*仲裁', r'申请.*仲裁', r'诉前告知',
                       r'最后.*还款', r'立即还款', r'拒不.*处理']
    for pat in collection_pats:
        if re.search(pat, text):
            return 'HARASS', '高', '匹配催收模式'
    
    # TRANSACTION
    # Government/public service
    gov_kws = ['中国残联', '助残日', '共青团', '12355', '公益热线', '防汛',
               '抗旱', '气象预警', '教育部', '整治.*办学', '严禁提前开学',
               '存款保险', '全国科技工作者日', '科协', '防汛抗旱指挥部',
               '铁路安全', '电力公司', '国网', '气象服务', '手机气象',
               '公益短信', '廉洁', '纪检', '纪委监委']
    is_gov = any(kw in text for kw in gov_kws)
    
    # Express/delivery
    express_kws = ['取件码', '凭.*领取', '运单尾号', '包裹.*派送',
                   '韵达', '顺丰', '中通', '圆通', '极兔', '京东配送',
                   '菜鸟驿站', '丰巢', '妈妈驿站', '已到.*取件',
                   '派送上门', '请.*取件', '包裹.*已到', '代收点']
    is_express = any(kw in text for kw in express_kws)
    
    # Operator service notifications (流量/余额/账单/验证码 etc.)
    operator_trans_kws = [
        '流量.*使用', '语音.*使用', '余额', '话费', '套餐', '月租',
        '充值', '消费', '积分.*领取', '权益.*领取',
        '日包', '月包', '流量包', '语音包', '叠加包',
        '流量.*到账', '语音.*到账', '电子券.*到账',
        '流量.*使用完', '语音.*使用完', '流量用尽', '语音用尽',
        '余额不足', '扣费', '代交', '代扣', '充值成功',
        '交费', '查询服务', '服务评价', '满意度',
        '余额变动', '话费余额', '账户余额',
        '自动充值', '自动充', '流量超套', '流量.*提醒',
        '业务办理', '确认办理', '服务.*提醒',
        '中国移动', '中国联通', '中国电信',
        '未接来电', '遇忙未接', '拒接主叫', '来电提醒',
        '验证码', '验证码为', '安全验证', '登录验证',
    ]
    operator_trans_count = sum(1 for kw in operator_trans_kws if kw in text)
    
    # Bank/financial transactions
    bank_trans_kws = [
        '还款', '扣款', '到账', '账单', '提现', '还款成功', '还款失败',
        '信用卡', '银行', '借呗', '花呗', '本期账单', '未还', '已还',
        '尾号', '卡号', '消费提醒', '支出提醒', '收入提醒',
        '退款', '返现', '返还', '到账通知',
        '信用卡还款', '账单已结清', '本期.*已还',
        '支付宝', '微信支付', '京东支付',
        '自动还款', '扣款成功', '扣款失败', '代扣.*成功',
        '消费.*元', '支出.*元', '收入.*元',
    ]
    bank_trans_count = sum(1 for kw in bank_trans_kws if kw in text)
    
    trans_kws = operator_trans_kws + bank_trans_kws
    trans_count = sum(1 for kw in trans_kws if kw in text)
    
    # Strong transaction patterns (specific amounts, codes, dates)
    strong_trans_pats = [
        r'还款.*\d+\.?\d*元', r'到账.*\d+\.?\d*元', r'余额.*\d+\.?\d*元',
        r'账单.*\d+\.?\d*元', r'消费.*\d+\.?\d*元', r'充值.*\d+\.?\d*元',
        r'尾号\w+.*\d+\.?\d*元', r'验证码.*\d{4,8}',
        r'\d+\.?\d*元.*还款', r'取件码.*\w+', r'凭\w+.*领取',
        r'\d+\.?\d*元.*到账', r'\d+\.?\d*元.*消费',
        r'支出\s*\d+\.?\d*元', r'收入\s*\d+\.?\d*元',
    ]
    strong_trans = any(re.search(pat, text) for pat in strong_trans_pats)
    
    # Insurance, securities transactions
    insurance_kws = ['投保', '保单', '保险', '理赔', '续保', '车险', '入账提醒']
    if any(kw in text for kw in insurance_kws):
        return 'TRANSACTION', '高', '保险/金融服务通知'
    
    # Vehicle service
    vehicle_kws = ['车险', '车辆', '车主.*授权', '车牌']
    if any(kw in text for kw in vehicle_kws):
        return 'TRANSACTION', '中', '车辆服务通知'
    
    # Government/public service (NEEDS_REVIEW)
    if is_gov and not any(kw in text for kw in ['权益', '会员', '流量包', '优惠', '红包', '福利', '活动']):
        return 'NEEDS_REVIEW', '高', '政府/公益/公共服务信息，无商业推广'
    
    # Express delivery (TRANSACTION)
    if is_express and not any(kw in text for kw in ['贷款', '额度', '借款', '审核', '权益', '会员']):
        return 'TRANSACTION', '高', '快递/物流服务通知'
    
    # Strong transaction with amount
    if strong_trans and trans_count >= 2:
        return 'TRANSACTION', '高', '包含明确的交易信息（金额/账单/到账/验证码）'
    
    # Operator services (流量/余额/账单/未接来电 etc.)
    if operator_trans_count >= 2:
        return 'TRANSACTION', '高', '运营商服务通知（流量/余额/账单/来电提醒）'
    
    # Bank transactions
    if bank_trans_count >= 2 and strong_trans:
        return 'TRANSACTION', '高', '银行/金融交易通知'
    
    # Work/business notifications
    work_kws = ['施工', '派单', '跳纤', '网元', '告警', '工单', '检修',
                '派送.*单', '售后单', '服务单']
    if any(kw in text for kw in work_kws):
        return 'NEEDS_REVIEW', '中', '工作/业务通知，需人工判断'
    
    # Weather
    weather_kws = ['今晚到明天', '明晚到后天', '多云', '晴', '气温',
                   '风力', '湿度', '摄氏度', '度']
    if any(kw in text for kw in weather_kws) and len(text) < 200:
        return 'NEEDS_REVIEW', '高', '天气预报服务信息'
    
    # Evaluation/survey invitations
    eval_kws = ['满意度', '服务评价', '评价.*抽奖', '评价.*参与', '调研',
                '满意度调查', '请您评价', '诚邀.*评价']
    if any(kw in text for kw in eval_kws):
        if any(kw in text for kw in ['抽奖', '赢', '礼品', '话费券', '激励']):
            return 'AD', '中', '服务评价邀请，含营销激励（抽奖/礼品）'
        else:
            return 'NEEDS_REVIEW', '中', '服务评价邀请，边界模糊'
    
    # AD (advertisements)
    loan_kws = ['借款.*已', '额度.*已', '贷款.*已', '预授信', '预放', '预审',
                '消费额', '信用贷', '激活.*额度', '查看.*额度', '领取.*额度',
                r'\d+元.*已', '预汇入', '预转入', '成功.*\d+元', '预放.*\d+',
                '审核通过.*\d+', '预批.*\d+', '循环可支用', '待激活',
                '待提款', '待提取', '额度待激活', '可支用.*元',
                '预估.*\d+万', '预估.*\d+元', '特批.*\d+', '放宽.*\d+',
                '预进入.*元', '预汇入.*元']
    is_loan_ad = any(re.search(kw, text) if kw.startswith('r') else kw in text for kw in loan_kws)
    if is_loan_ad:
        return 'AD', '高', '贷款产品推广广告'
    
    ad_kws = ['贷款', '额度', '借款', '点击.*领取', '点击.*查看', '点击.*查询',
              '点击.*参与', '拒收请回复R', '拒收请回复 R', '会员', '权益',
              '优惠', '红包', '福利', '限时', '免费领取', '免费', '特价',
              '恭喜.*获得', '预授信', '预批', '特批', '放宽', '咨询电话',
              '客服电话', '详询.*\d{3,}', 'http', '点击.*报名',
              '点击.*确认', '点击.*激活', '点击.*提取', '点击.*参与',
              '报名.*活动', '赢取.*大奖', '赢取.*话费', '赢取.*礼品',
              '立减', '折上折', '预售', '上市', '新品']
    ad_count = sum(1 for kw in ad_kws if kw in text)
    
    # Medical ads
    medical_ad_kws = ['筛查', '普查', '补助', '援助', '申请.*回', '白癜风',
                      '银屑病', '胎记', '癫痫', '了解详情', '申请检查',
                      '惠民', '公益普查']
    is_medical_ad = any(kw in text for kw in medical_ad_kws)
    if is_medical_ad:
        return 'AD', '中', '医疗推广信息'
    
    # Education ads
    edu_ad_kws = ['课程', '培训', '资料已发', '教程已发', r'不点.*失效', '速领',
                  '涨分资料', '剪辑推广', '变现教程', '学习资料']
    is_edu_ad = any(kw in text for kw in edu_ad_kws)
    if is_edu_ad:
        return 'AD', '中', '教育培训推广'
    
    # Operator marketing
    operator_kws = ['中国移动', '中国联通', '中国电信']
    is_operator = any(kw in text for kw in operator_kws)
    marketing_kws = ['权益', '会员', '优惠', '红包', '福利', '限时', '免费领取',
                     '恭喜.*获得', '点击.*领取', '点击.*参与', '抽奖', '赢取',
                     '活动.*开启', '活动.*上线', '嘉年华', '礼遇']
    if is_operator and any(kw in text for kw in marketing_kws):
        return 'AD', '中', '运营商营销推广'
    
    # Product ads
    product_ad_kws = ['预售', '上市', '新品', '抢购', '秒杀', '特惠', '折扣',
                      '满减', '立减', '折上折']
    if any(kw in text for kw in product_ad_kws) and ad_count >= 2:
        return 'AD', '高', '产品营销广告'
    
    if ad_count >= 3:
        return 'AD', '高', '包含多个广告特征'
    
    # NEEDS_REVIEW
    if is_gov:
        return 'NEEDS_REVIEW', '高', '政府/公益信息'
    
    if trans_count >= 1:
        return 'TRANSACTION', '中', '疑似交易/服务通知，需人工确认'
    
    return 'NEEDS_REVIEW', '低', '无法明确归类，需人工判断'


# Load workbook
print('Loading workbook...')
wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
print(f'Total rows: {len(rows)}')

# Add header if needed
header = rows[0]
if len(header) < 7:
    ws.cell(row=1, column=5, value='final_label')
    ws.cell(row=1, column=6, value='confidence')
    ws.cell(row=1, column=7, value='rationale')

# Classify and write
results = []
counts = Counter()
for i in range(1, len(rows)):
    text = rows[i][0] if rows[i][0] else ''
    orig_label = rows[i][1] if len(rows[i]) > 1 else ''
    label, conf, reason = classify_row(text, orig_label)
    
    ws.cell(row=i+1, column=5, value=label)
    ws.cell(row=i+1, column=6, value=conf)
    ws.cell(row=i+1, column=7, value=reason)
    
    results.append((i+1, label, conf, reason))
    counts[label] += 1

# Save
print('Saving workbook...')
wb.save(INPUT_FILE)
print('Done!')

# Write CSV
print('Writing CSV...')
with open(OUTPUT_CSV, 'w', encoding='utf-8') as f:
    f.write('row_num,final_label,confidence,rationale\n')
    for row_num, label, conf, reason in results:
        reason_escaped = reason.replace('"', "'").replace('\n', ' ')
        f.write(f'{row_num},{label},{conf},"{reason_escaped}"\n')

print('\n=== Annotation Summary ===')
total = len(results)
for label in ['TRANSACTION', 'AD', 'HARASS', 'FRAUD', 'NEEDS_REVIEW']:
    count = counts.get(label, 0)
    pct = count / total * 100 if total > 0 else 0
    print(f'{label}: {count} ({pct:.1f}%)')
print(f'Total: {total}')
print(f'\nCSV saved: {OUTPUT_CSV}')
