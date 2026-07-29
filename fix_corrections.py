# -*- coding: utf-8 -*-
"""Fix manual corrections for edge cases"""
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

INPUT_FILE = r'C:\Users\woshinibaba\Downloads\normal_2w_output.xlsx'

# Manual corrections based on detailed review
CORRECTIONS = {
    # Row 4: 联通Assistant未接来电 → AD
    4: ('AD', '高', '联通Assistant营销推送，推广5G通信服务与微信公众号'),
    
    # Row 5: 淘宝评价通知（含现金券激励）→ AD
    5: ('AD', '高', '电商平台评价邀请，含现金券激励，属营销推广'),
    
    # Row 11: EMS快递取件 → TRANSACTION
    11: ('TRANSACTION', '高', 'EMS快递取件通知，含取件码与地址信息'),
    
    # Row 23: 催收信息（划扣+12点前）→ HARASS
    23: ('HARASS', '高', '催收信息，称业务已到期系统将划扣，要求12点前登录app处理'),
    
    # Row 31: 联通生日祝福+积分兑换 → AD
    31: ('AD', '高', '联通生日营销，推送生日积分兑换视频VIP等礼品，属运营商营销'),
    
    # Row 40: 联通Assistant → AD
    40: ('AD', '高', '联通Assistant营销推送，推广5G通信服务与微信公众号'),
    
    # Row 49: 快手余额通知（含提现营销）→ AD
    49: ('AD', '高', '快手账户余额通知，含今日登录提现营销引导，属平台推广'),
    
    # Row 57: 消费券抵扣（588000余额可疑）→ AD
    57: ('AD', '高', '消费券抵扣提醒，含588000余额诱导点击，属营销广告'),
    
    # Row 59: 蚂蚁智信花呗欠款 → HARASS
    59: ('HARASS', '高', '花呗欠款催收，称至今未处理，引导登录支付宝还款'),
    
    # Row 63: 评价赢话费券 → AD
    63: ('AD', '高', '中国移动满意度调研邀请，含3元话费券激励，属营销推广'),
    
    # Row 64: 京东服务单取件 → TRANSACTION
    64: ('TRANSACTION', '高', '京东服务单预约取件通知，告知将上门取件'),
    
    # Row 70: 服务评价邀请 → NEEDS_REVIEW
    70: ('NEEDS_REVIEW', '中', '中国移动服务评价邀请，边界模糊'),
    
    # Row 87: 农业银行授信 → AD
    87: ('AD', '高', '农业银行贷款广告，称预授信328000元，引导回复办理'),
    
    # Row 91: 中国银行授信 → AD
    91: ('AD', '高', '中国银行贷款广告，称预授信328000元，引导回复办理'),
    
    # Row 95: 借呗利率降低 → AD
    95: ('AD', '高', '借呗营销推广，称利率已降低至年化14.76%，引导点击使用'),
    
    # Row 96: 全民钱包 → AD
    96: ('AD', '高', '贷款广告，称今日已向您账户成功预进入68876.5元'),
    
    # Row 51: 星火保入账 → TRANSACTION
    51: ('TRANSACTION', '高', '星火保补贴入账提醒，属保险金融服务通知'),
    
    # Row 52: 比亚迪车辆授权 → TRANSACTION
    52: ('TRANSACTION', '高', '比亚迪车辆授权通知，告知可使用车辆驱逐舰05'),
}

print('Loading workbook...')
wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb.active

corrected = 0
for row_num, (label, conf, reason) in CORRECTIONS.items():
    ws.cell(row=row_num, column=5, value=label)
    ws.cell(row=row_num, column=6, value=conf)
    ws.cell(row=row_num, column=7, value=reason)
    corrected += 1

print(f'Corrected {corrected} rows')
print('Saving workbook...')
wb.save(INPUT_FILE)
print('Done!')

# Verify corrections
rows = list(ws.iter_rows(values_only=True))
print('\n=== Corrected Samples ===')
for i in [4, 5, 11, 23, 31, 40, 49, 57, 59, 63, 64, 70, 87, 91, 95, 96, 51, 52]:
    text = str(rows[i-1][0])[:60] if rows[i-1][0] else ''
    label = rows[i-1][4] if len(rows[i-1]) > 4 else ''
    print(f'{i}: [{label}] {text}')
