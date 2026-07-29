# -*- coding: utf-8 -*-
"""Generate complete annotations for all 10000 SMS rows"""
import openpyxl, sys, re
from collections import Counter
sys.stdout.reconfigure(encoding='utf-8')

INPUT_FILE = r'C:\Users\woshinibaba\Downloads\normal_2w_output.xlsx'
OUTPUT_CSV = r'C:\Users\woshinibaba\Documents\oppo的项目\Android_SMS_Classifier\annotations_all_10000.csv'

def classify_row(text, orig_label):
    if not text or not str(text).strip():
        return 'NEEDS_REVIEW', '低', '短信内容为空'
    text = str(text)
    
    # FRAUD
    fraud_kws = ['涉嫌诈骗', '涉嫌恶意透支', '冻结支付账户', '法院传票',
                 '拘捕令', '涉嫌洗钱', '包裹涉毒', '社保异常', '逮捕令']
    if any(kw in text for kw in fraud_kws):
        return 'FRAUD', '高', '包含诈骗关键词'
    
    # HARASS
    harass_kws = ['严重逾期', '恶意透支', '强制执行', '起诉', '法庭',
                  '上报征信', '列入失信', '冻结账户', '划扣工资', '联系单位',
                  '户籍地', '催收', '催缴', '逾期记录', '报送人行',
                  '征信系统', '不良信用', '默认拒偿', '永不减免',
                  '申请司法', '诉讼', '开庭', '最后期限', '否则将']
    if sum(1 for kw in harass_kws if kw in text) >= 2:
        return 'HARASS', '高', '包含多个催收/威胁关键词'
    
    collection_pats = [r'合同逾期.*全款偿还', r'分期资格.*取消',
                       r'强制执行.*查封', r'催缴.*公开', r'联系.*亲属']
    for pat in collection_pats:
        if re.search(pat, text):
            return 'HARASS', '高', '匹配催收模式'
    
    # TRANSACTION
    gov_kws = ['中国残联', '助残日', '共青团', '12355', '公益热线', '防汛',
               '抗旱', '气象预警', '教育部', '整治.*办学', '严禁提前开学',
               '存款保险', '全国科技工作者日', '科协', '防汛抗旱指挥部']
    is_gov = any(kw in text for kw in gov_kws)
    
    express_kws = ['取件码', '凭.*领取', '运单尾号', '包裹.*派送',
                   '韵达', '顺丰', '中通', '圆通', '极兔', '京东配送',
                   '菜鸟驿站', '丰巢', '妈妈驿站', '已到.*取件']
    is_express = any(kw in text for kw in express_kws)
    
    trans_kws = ['还款', '扣款', '到账', '余额', '账单', '提现', '充值', '消费',
                 '收入', '支出', '转账', '汇款', '入账', '自动还款', '还款成功',
                 '还款失败', '信用卡', '银行', '借呗', '花呗', '信用卡账单',
                 '本期账单', '未还', '已还', '尾号', '卡号', '消费提醒', '积分',
                 '券', '红包', '电子券', '流量', '话费', '退款', '返现', '返还',
                 '订单', '派送', '已发货', '已送达', '配送', '物流', '验证码',
                 '验证码为', '余额不足', '扣费', '代交', '代扣', '充值成功',
                 '交费', '查询服务', '服务评价', '满意度', '话费.*查询',
                 '账单.*查询', '流量.*查询', '余额.*查询', '自动充值',
                 '余额变动', '话费余额', '账户余额', '积分.*领取',
                 '权益.*领取', '月卡权益', '日包', '月包', '套餐']
    trans_count = sum(1 for kw in trans_kws if kw in text)
    
    strong_trans_pats = [r'还款.*\d+\.?\d*元', r'到账.*\d+\.?\d*元',
                         r'余额.*\d+\.?\d*元', r'账单.*\d+\.?\d*元',
                         r'消费.*\d+\.?\d*元', r'充值.*\d+\.?\d*元',
                         r'尾号\w+.*\d+\.?\d*元', r'验证码.*\d{4,8}',
                         r'\d+\.?\d*元.*还款', r'取件码.*\w+',
                         r'凭\w+.*领取', r'尾号.*\d+']
    strong_trans = any(re.search(pat, text) for pat in strong_trans_pats)
    
    if is_gov and not any(kw in text for kw in ['权益', '会员', '流量', '优惠', '红包', '福利']):
        return 'NEEDS_REVIEW', '高', '政府/公益信息，无商业推广'
    
    if is_express and not any(kw in text for kw in ['贷款', '额度', '借款', '审核']):
        return 'TRANSACTION', '高', '快递/物流服务通知'
    
    if strong_trans and trans_count >= 2:
        return 'TRANSACTION', '高', '包含明确的交易信息（金额/账单/到账）'
    
    if trans_count >= 4:
        return 'TRANSACTION', '高', '运营商服务通知（余额/流量/账单）'
    
    # AD
    loan_kws = ['借款.*已', '额度.*已', '贷款.*已', '预授信', '预放', '预审',
                '消费额', '信用贷', '激活.*额度', '查看.*额度', '领取.*额度',
                r'\d+元.*已', '预汇入', '预转入', '成功.*\d+元', '预放.*\d+',
                '审核通过.*\d+', '预批.*\d+', '循环可支用', '待激活',
                '待提款', '待提取', '额度待激活']
    is_loan_ad = any(re.search(kw, text) if kw.startswith('r') else kw in text for kw in loan_kws)
    if is_loan_ad:
        return 'AD', '高', '贷款产品推广广告'
    
    ad_kws = ['贷款', '额度', '借款', '点击.*领取', '点击.*查看', '点击.*查询',
              '点击.*参与', '拒收请回复R', '拒收请回复 R', '会员', '权益',
              '优惠', '红包', '福利', '限时', '免费领取', '免费', '特价',
              '恭喜.*获得', '预授信', '预批', '特批', '放宽', '咨询电话',
              '客服电话', '详询.*\d{3,}', 'http', '点击.*报名',
              '点击.*确认', '点击.*激活', '点击.*提取', '点击.*参与']
    ad_count = sum(1 for kw in ad_kws if kw in text)
    
    medical_ad_kws = ['筛查', '普查', '补助', '援助', '申请.*回', '白癜风',
                      '银屑病', '胎记', '癫痫', '了解详情', '申请检查']
    is_medical_ad = any(kw in text for kw in medical_ad_kws)
    if is_medical_ad:
        return 'AD', '中', '医疗推广信息'
    
    edu_ad_kws = ['课程', '培训', '资料已发', '教程已发', r'不点.*失效', '速领',
                  '涨分资料', '剪辑推广', '变现教程']
    is_edu_ad = any(kw in text for kw in edu_ad_kws)
    if is_edu_ad:
        return 'AD', '中', '教育培训推广'
    
    operator_kws = ['中国移动', '中国联通', '中国电信']
    is_operator = any(kw in text for kw in operator_kws)
    marketing_kws = ['权益', '会员', '优惠', '红包', '福利', '限时', '免费领取',
                     '恭喜.*获得', '点击.*领取', '点击.*参与', '抽奖', '赢取']
    if is_operator and any(kw in text for kw in marketing_kws):
        return 'AD', '中', '运营商营销推广'
    
    if ad_count >= 3:
        return 'AD', '高', '包含多个广告特征'
    
    # NEEDS_REVIEW
    news_kws = ['江西日报', '央视新闻', '新华社', '中新网', '人民日报',
                '大江网', '信息日报', '四川手机报', '山西日报', '新闻', '日讯']
    is_news = any(kw in text for kw in news_kws)
    if is_news and ad_count >= 2:
        return 'NEEDS_REVIEW', '中', '新闻媒体信息，但含推广链接'
    
    personal_kws = ['您好', '尊敬的', '温馨提示', '会议.*通知', '订货单']
    is_personal = any(kw in text for kw in personal_kws)
    if is_personal and not is_express and trans_count < 2:
        return 'NEEDS_REVIEW', '低', '个人/商务通知，需人工判断'
    
    return 'NEEDS_REVIEW', '低', '无法明确归类，需人工判断'


# Load workbook
print('Loading workbook...')
wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
print(f'Total rows: {len(rows)}')

# Add header if needed
header = rows[0]
if len(header) < 7:
    ws.cell(row=1, column=5, value='final_label')
    ws.cell(row=1, column=6, value='confidence')
    ws.cell(row=1, column=7, value='rationale')

# Classify and write
results = []
counts = Counter()
for i in range(1, len(rows)):
    text = rows[i][0] if rows[i][0] else ''
    orig_label = rows[i][1] if len(rows[i]) > 1 else ''
    label, conf, reason = classify_row(text, orig_label)
    
    ws.cell(row=i+1, column=5, value=label)
    ws.cell(row=i+1, column=6, value=conf)
    ws.cell(row=i+1, column=7, value=reason)
    
    results.append((i+1, label, conf, reason))
    counts[label] += 1

# Save
print('Saving workbook...')
wb.save(INPUT_FILE)
print('Done!')

# Write CSV
print('Writing CSV...')
with open(OUTPUT_CSV, 'w', encoding='utf-8') as f:
    f.write('row_num,final_label,confidence,rationale\n')
    for row_num, label, conf, reason in results:
        reason_escaped = reason.replace('"', "'").replace('\n', ' ')
        f.write(f'{row_num},{label},{conf},"{reason_escaped}"\n')

print('\n=== Annotation Summary ===')
total = len(results)
for label in ['TRANSACTION', 'AD', 'HARASS', 'FRAUD', 'NEEDS_REVIEW']:
    count = counts.get(label, 0)
    pct = count / total * 100 if total > 0 else 0
    print(f'{label}: {count} ({pct:.1f}%)')
print(f'Total: {total}')
print(f'\nCSV saved: {OUTPUT_CSV}')
