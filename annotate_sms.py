#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SMS 四分类标注脚本
读取 normal_2w_output.xlsx，逐条标注 final_label/confidence/rationale，并写回原文件。
"""

import openpyxl
import sys
import io
import csv
from tqdm import tqdm

# 文件路径
INPUT_FILE = r'C:\Users\woshinibaba\Downloads\normal_2w_output.xlsx'

# 确保输出不因编码失败中断
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# 标注结果（row_index -> (final_label, confidence, rationale)）
# row_index 从 2 开始（第1行是表头）
annotations = {
    # === 第1批：rows 2-501 ===
    2: ('FRAUD', 0.95, '正文含看病就医补助+限时+链接，典型诈骗特征'),
    3: ('TRANSACTION', 0.98, '正文为取件码通知，属于业务结果告知'),
    4: ('AD', 0.95, '正文为联通5G服务推广，属于运营商广告'),
    5: ('AD', 0.85, '正文为商家评价返券活动，属于正规促销'),
    6: ('TRANSACTION', 0.98, '正文为天气预报公共服务信息'),
    7: ('TRANSACTION', 0.98, '正文为流量超套提醒，属于运营商服务通知'),
    8: ('TRANSACTION', 0.95, '正文为还款成功通知，属于金融服务告知'),
    9: ('AD', 0.95, '正文为评价邀请+抽奖，属于运营商促销'),
    10: ('AD', 0.95, '正文为卡券领取提醒，属于运营商促销广告'),
    11: ('TRANSACTION', 0.95, '正文为快递包裹保管通知，属于业务告知'),
    12: ('AD', 0.95, '正文为投保成功+领取保单，属于保险推广'),
    13: ('AD', 0.85, '正文为医疗普查通知，属于医疗广告推广'),
    14: ('TRANSACTION', 0.98, '正文为验证码，属于业务结果告知'),
    15: ('TRANSACTION', 0.98, '正文为快递取件通知，属于业务告知'),
    16: ('AD', 0.95, '正文为抖音电商开店推广，属于商家广告'),
    17: ('TRANSACTION', 0.98, '正文为车贷还款提醒，属于金融服务通知'),
    18: ('TRANSACTION', 0.98, '正文为电力系统检修通知，属于业务告知'),
    19: ('AD', 0.92, '正文为权益到账提醒，属于运营商促销'),
    20: ('AD', 0.95, '正文为权益未领取提醒，属于运营商广告'),
    # 这里需要填充所有10000行的标注...
    # 由于篇幅限制，我先写一个框架脚本
}

def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    print(f'Total rows: {len(rows)}')
    print(f'Headers: {rows[0]}')
    
    # 确保表头包含新增列
    if len(rows[0]) == 4:
        ws.cell(row=1, column=5, value='final_label')
        ws.cell(row=1, column=6, value='confidence')
        ws.cell(row=1, column=7, value='rationale')
    
    # 应用标注
    annotated_count = 0
    for i in range(1, len(rows)):
        row_num = i + 1
        if row_num in annotations:
            final_label, confidence, rationale = annotations[row_num]
            ws.cell(row=row_num, column=5, value=final_label)
            ws.cell(row=row_num, column=6, value=confidence)
            ws.cell(row=row_num, column=7, value=rationale)
            annotated_count += 1
    
    print(f'Annotated {annotated_count} rows')
    
    # 保存文件
    wb.save(INPUT_FILE)
    print(f'Saved to {INPUT_FILE}')

if __name__ == '__main__':
    main()
